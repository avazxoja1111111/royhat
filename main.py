import logging
import asyncio
import json
import os
import random
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import tempfile

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, 
    InlineKeyboardButton, BufferedInputFile
)
from aiogram.enums import ParseMode
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.client.default import DefaultBotProperties

# PDF and Excel libraries
try:
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import PyPDF2
    from io import BytesIO
except ImportError:
    print("Required libraries not installed. Install with: pip install reportlab openpyxl PyPDF2")

# 🔑 Configuration
TOKEN = os.getenv("BOT_TOKEN", "7570796885:AAHHfpXanemNYvW-wVT2Rv40U0xq-XjxSwk")
SUPER_ADMIN_ID = int(os.getenv("SUPER_ADMIN_ID", "6578706277, 7853664401"))
CHANNEL_USERNAME = "@Kitobxon_Kids"

# 🛠 Logging
logging.basicConfig(level=logging.INFO)

# 🤖 Bot and Dispatcher
bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

# 📁 Data storage files
DATA_DIR = "bot_data"
os.makedirs(DATA_DIR, exist_ok=True)

USERS_FILE = os.path.join(DATA_DIR, "users.json")
ADMINS_FILE = os.path.join(DATA_DIR, "admins.json")
TESTS_FILE = os.path.join(DATA_DIR, "tests.json")
RESULTS_FILE = os.path.join(DATA_DIR, "results.json")

# 📌 Regions data
REGIONS = {
    "Toshkent shahri": ["Bektemir", "Chilonzor", "Mirzo Ulug'bek", "Mirobod", "Olmazor", "Shayxontohur", "Sergeli", "Uchtepa", "Yashnobod", "Yakkasaroy", "Yunusobod"],
    "Toshkent viloyati": ["Bekabad", "Bo'ka", "Bo'stonliq", "Chinoz", "Chirchiq", "Ohangaron", "Oqqo'rg'on", "Parkent", "Piskent", "Quyichirchiq", "O'rtachirchiq", "Yangiyo'l", "Toshkent", "Yuqorichirchiq", "Zangiota", "Nurafshon", "Olmaliq", "Angren"],
    "Andijon": ["Andijon shahri", "Asaka", "Baliqchi", "Bo'ston", "Buloqboshi", "Izboskan", "Jalaquduq", "Marhamat", "Oltinko'l", "Paxtaobod", "Paytug'", "Qo'rg'ontepa", "Shahriston", "Xo'jaobod"],
    "Farg'ona": ["Beshariq", "Buvayda", "Dang'ara", "Farg'ona shahri", "Ferghana tumani", "Furqat", "Qo'qon", "Quva", "Rishton", "So'x", "Toshloq", "Uchko'prik", "Yozyovon", "Oltiariq"],
    "Namangan": ["Chortoq", "Chust", "Kosonsoy", "Namangan shahri", "Norin", "Pop", "To'raqo'rg'on", "Uychi", "Uchqo'rg'on", "Yangiqo'rg'on", "Yangihayot"],
    "Samarqand": ["Bulung'ur", "Ishtixon", "Jomboy", "Kattakurgan", "Oqdaryo", "Payariq", "Pastdarg'om", "Qo'shrabot", "Samarqand shahri", "Toyloq", "Urgut"],
    "Buxoro": ["Buxoro shahri", "Buxoro tumani", "G'ijduvon", "Jondor", "Kogon", "Olot", "Peshku", "Qorako'l", "Qorovulbozor", "Romitan", "Shofirkon", "Vobkent"],
    "Jizzax": ["Baxmal", "Chiroqchi", "Do'stlik", "Forish", "G'allaorol", "Zafarobod", "Zarbdor", "Zomin", "Zafar", "Yangiobod", "Jizzax shahri", "Mirzacho'l"],
    "Navoiy": ["Bespah", "Karmana", "Konimex", "Navbahor", "Nurota", "Tomdi", "Xatirchi", "Uchquduq", "Navoiy shahri", "Zarafshon"],
    "Qashqadaryo": ["Chiroqchi", "G'uzor", "Qarshi", "Kitob", "Koson", "Mirishkor", "Muborak", "Nishon", "Shahrisabz", "Dehqonobod", "Yakkabog'"],
    "Surxondaryo": ["Angor", "Bandixon", "Denov", "Jarqo'rg'on", "Muzrabot", "Oltinsoy", "Sariosiyo", "Sherobod", "Sho'rchi", "Termiz", "Uzun", "Boysun"],
    "Sirdaryo": ["Guliston", "Guliston tumani", "Mirzaobod", "Oqoltin", "Sardoba", "Sayxunobod", "Sirdaryo tumani", "Xovos", "Boyovut", "Yangiyer"],
    "Xorazm": ["Bog'ot", "Gurlan", "Hazorasp", "Khiva", "Qo'shko'pir", "Shovot", "Urganch tumani", "Xonqa", "Yangiariq", "Yangibozor", "Tuproqqal'a", "Urganch shahri"],
    "Qoraqalpog'iston": ["Amudaryo", "Beruniy", "Chimboy", "Ellikqala", "Kegeyli", "Mo'ynoq", "Nukus", "Qanliko'l", "Qo'ng'irot", "Taxiatosh", "To'rtko'l", "Xo'jayli"]
}

# 📌 Data management functions
def load_json_data(file_path: str, default_data: Any = None) -> Any:
    """Load data from JSON file"""
    if not os.path.exists(file_path):
        return default_data or {}
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        return default_data or {}

def save_json_data(file_path: str, data: Any) -> None:
    """Save data to JSON file"""
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_users() -> Dict:
    """Get all registered users"""
    return load_json_data(USERS_FILE, {})

def save_user(user_id: str, user_data: Dict) -> None:
    """Save user data"""
    users = get_users()
    users[user_id] = user_data
    save_json_data(USERS_FILE, users)

def get_admins() -> Dict:
    """Get all admins"""
    default_admins = {str(SUPER_ADMIN_ID): {"role": "super_admin", "added_by": "system", "added_date": datetime.now().isoformat()}}
    return load_json_data(ADMINS_FILE, default_admins)

def save_admin(admin_id: str, admin_data: Dict) -> None:
    """Save admin data"""
    admins = get_admins()
    admins[admin_id] = admin_data
    save_json_data(ADMINS_FILE, admins)

def remove_admin(admin_id: str) -> bool:
    """Remove admin by ID"""
    try:
        admins = get_admins()
        if admin_id in admins:
            del admins[admin_id]
            save_json_data(ADMINS_FILE, admins)
            return True
        return False
    except Exception as e:
        logging.error(f"Error removing admin data: {e}")
        return False

def get_tests() -> Dict:
    """Get all tests"""
    return load_json_data(TESTS_FILE, {"7-10": {}, "11-14": {}})

def save_test(test_data: Dict) -> None:
    """Save test data"""
    tests = get_tests()
    age_group = test_data["age_group"]
    test_id = str(uuid.uuid4())
    
    if age_group not in tests:
        tests[age_group] = {}
    
    tests[age_group][test_id] = test_data
    save_json_data(TESTS_FILE, tests)

def get_results() -> List:
    """Get all test results"""
    return load_json_data(RESULTS_FILE, [])

def save_result(result_data: Dict) -> None:
    """Save test result"""
    results = get_results()
    results.append(result_data)
    save_json_data(RESULTS_FILE, results)

def is_admin(user_id: int) -> bool:
    """Check if user is admin"""
    admins = get_admins()
    return str(user_id) in admins

def is_super_admin(user_id: int) -> bool:
    """Check if user is super admin"""
    admins = get_admins()
    return str(user_id) in admins and admins[str(user_id)].get("role") == "super_admin"

# 📌 FSM States
class Registration(StatesGroup):
    check_subscription = State()
    child_name = State()
    parent_name = State()
    region = State()
    district = State()
    mahalla = State()
    age = State()
    phone = State()
    feedback = State()

class AdminStates(StatesGroup):
    add_admin = State()
    remove_admin = State()
    promote_super_admin = State()
    add_test_age = State()
    add_test_book = State()
    add_test_content = State()
    add_test_questions = State()
    delete_test_age = State()
    delete_test_select = State()

class TestStates(StatesGroup):
    taking_test = State()
    test_question = State()

# 📌 Keyboards
def get_main_menu():
    """Main menu keyboard"""
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📋 Ro'yxatdan o'tish")],
        [KeyboardButton(text="📝 Test topshirish")],
        [KeyboardButton(text="💬 Fikr va maslahatlar")],
        [KeyboardButton(text="📚 Loyiha haqida")]
    ], resize_keyboard=True)

def get_admin_menu(is_super: bool = False):
    """Admin menu keyboard"""
    keyboard = [
        [KeyboardButton(text="👥 Foydalanuvchilar ro'yxati")],
        [KeyboardButton(text="➕ Test qo'shish")]
    ]
    
    if is_super:
        keyboard.extend([
            [KeyboardButton(text="👨‍💼 Adminlar ro'yxati")],
            [KeyboardButton(text="➕ Admin qo'shish")],
            [KeyboardButton(text="⬆️ Super Admin tayinlash")],
            [KeyboardButton(text="➖ Admin o'chirish")],
            [KeyboardButton(text="🗑 Test o'chirish")],
            [KeyboardButton(text="📊 Test natijalarini yuklab olish")],
            [KeyboardButton(text="📋 Foydalanuvchi ma'lumotlarini yuklab olish")]
        ])
    
    keyboard.append([KeyboardButton(text="🔙 Asosiy menyu")])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)

def get_age_group_keyboard():
    """Age group selection keyboard"""
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="7-10 yosh")],
        [KeyboardButton(text="11-14 yosh")],
        [KeyboardButton(text="🔙 Orqaga")]
    ], resize_keyboard=True)

# 📌 PDF and Excel generation functions
def generate_pdf_report(results: List[Dict]) -> bytes:
    """Generate PDF report of test results"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("KITOBXON KIDS - Test Natijalari", styles['Title'])
    
    # Create table data
    table_data = [['Foydalanuvchi', 'Yosh', 'Ball', 'Vaqt', 'Foiz', 'Sana']]
    
    for result in results:
        table_data.append([
            result.get('user_name', 'N/A'),
            result.get('age', 'N/A'),
            f"{result.get('score', 0)}/100",
            result.get('time_taken', 'N/A'),
            f"{result.get('percentage', 0)}%",
            result.get('date', 'N/A')
        ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    # Build PDF
    doc.build([title, table])
    buffer.seek(0)
    return buffer.getvalue()

def generate_users_pdf_report(users_data: Dict) -> bytes:
    """Generate PDF report of registered users"""
    buffer = BytesIO()
    # Use landscape orientation and larger page size for more space
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("KITOBXON KIDS - Ro'yxatdan o'tgan foydalanuvchilar", styles['Title'])
    
    # Create table data with wrapped text
    table_data = [['Farzand nomi', 'Ota-ona', 'Yosh', 'Viloyat', 'Tuman', 'Telegram ID', 'Username', 'Telefon', 'Ro\'yxat sanasi']]
    
    for user_id, user_data in users_data.items():
        username = user_data.get('username', 'N/A')
        if username != 'N/A' and not username.startswith('@'):
            username = f"@{username}"
        
        # Wrap long text in Paragraphs for better display
        child_name = Paragraph(user_data.get('child_name', 'N/A'), styles['Normal'])
        parent_name = Paragraph(user_data.get('parent_name', 'N/A'), styles['Normal'])
        region = Paragraph(user_data.get('region', 'N/A'), styles['Normal'])
        district = Paragraph(user_data.get('district', 'N/A'), styles['Normal'])
        username_p = Paragraph(username, styles['Normal'])
        phone = Paragraph(user_data.get('phone', 'N/A'), styles['Normal'])
        
        table_data.append([
            child_name,
            parent_name,
            user_data.get('age', 'N/A'),
            region,
            district,
            user_id,
            username_p,
            phone,
            user_data.get('registration_date', 'N/A')[:10] if user_data.get('registration_date') else 'N/A'
        ])
    
    # Create table with specific column widths
    col_widths = [80, 80, 35, 70, 70, 60, 70, 70, 65]  # Adjusted widths
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.lightgrey]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    # Build PDF
    doc.build([title, table])
    buffer.seek(0)
    return buffer.getvalue()

def generate_users_excel_report(users_data: Dict) -> bytes:
    """Generate Excel report of registered users"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    
    # Headers
    headers = ['Farzand nomi', 'Ota-ona', 'Yosh', 'Viloyat', 'Tuman', 'Mahalla', 'Telegram ID', 'Username', 'Telefon', 'Ro\'yxat sanasi']
    ws.append(headers)
    
    # Style headers and set column widths
    column_widths = [20, 25, 8, 18, 18, 20, 15, 18, 18, 15]
    for i, (cell, width) in enumerate(zip(ws[1], column_widths), 1):
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        
        # Set column width
        column_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[column_letter].width = width
    
    # Add data with proper formatting
    for user_id, user_data in users_data.items():
        username = user_data.get('username', 'N/A')
        if username != 'N/A' and not username.startswith('@'):
            username = f"@{username}"
            
        row_data = [
            user_data.get('child_name', 'N/A'),
            user_data.get('parent_name', 'N/A'),
            user_data.get('age', 'N/A'),
            user_data.get('region', 'N/A'),
            user_data.get('district', 'N/A'),
            user_data.get('mahalla', 'N/A'),
            user_id,
            username,
            user_data.get('phone', 'N/A'),
            user_data.get('registration_date', 'N/A')[:10] if user_data.get('registration_date') else 'N/A'
        ]
        
        ws.append(row_data)
        
        # Format the newly added row
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(size=10)
            
            # Add alternating row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Set row heights for better visibility
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 25
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_excel_report(results: List[Dict]) -> bytes:
    """Generate Excel report of test results"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Natijalari"
    
    # Headers
    headers = ['Foydalanuvchi', 'Telegram ID', 'Username', 'Yosh', 'Ball', 'Vaqt', 'Foiz', 'Sana']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for result in results:
        ws.append([
            result.get('user_name', 'N/A'),
            result.get('telegram_id', 'N/A'),
            result.get('username', 'N/A'),
            result.get('age', 'N/A'),
            f"{result.get('score', 0)}/100",
            result.get('time_taken', 'N/A'),
            f"{result.get('percentage', 0)}%",
            result.get('date', 'N/A')
        ])
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """Extract text from PDF file"""
    try:
        pdf_reader = PyPDF2.PdfReader(BytesIO(pdf_bytes))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        logging.error(f"Error extracting text from PDF: {e}")
        return ""

# 📌 Main bot handlers

@dp.message(Command("start"))
async def start(message: types.Message, state: FSMContext):
    """Start command handler"""
    user_id = message.from_user.id
    
    # Check subscription
    try:
        chat_member = await bot.get_chat_member(chat_id=CHANNEL_USERNAME, user_id=user_id)
        if chat_member.status not in ["member", "administrator", "creator"]:
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✉️ Obuna bo'lish", url=f"https://t.me/{CHANNEL_USERNAME[1:]}")],
                [InlineKeyboardButton(text="✅ Obuna bo'ldim", callback_data="check_sub")]
            ])
            await message.answer("✉️ Iltimos, quyidagi kanalga obuna bo'ling:", reply_markup=keyboard)
            await state.set_state(Registration.check_subscription)
            return
    except Exception as e:
        logging.error(f"Error checking subscription: {e}")
    
    # Check if user is admin
    if is_admin(user_id):
        is_super = is_super_admin(user_id)
        role = "Super Admin" if is_super else "Admin"
        await message.answer(f"👋 Salom, {role}! Admin paneliga xush kelibsiz!", 
                           reply_markup=get_admin_menu(is_super))
    else:
        await message.answer("👋 Salom! 'KITOBXON KIDS' botiga xush kelibsiz!", 
                           reply_markup=get_main_menu())

@dp.callback_query(F.data == "check_sub")
async def check_subscription(callback_query: types.CallbackQuery, state: FSMContext):
    """Check subscription callback"""
    user_id = callback_query.from_user.id
    
    try:
        chat_member = await bot.get_chat_member(chat_id=CHANNEL_USERNAME, user_id=user_id)
        if chat_member.status not in ["member", "administrator", "creator"]:
            await callback_query.answer("❌ Hali ham obuna emassiz!", show_alert=True)
            return
    except Exception as e:
        logging.error(f"Error checking subscription: {e}")
    
    if is_admin(user_id):
        is_super = is_super_admin(user_id)
        role = "Super Admin" if is_super else "Admin"
        await bot.send_message(user_id, f"👋 Salom, {role}! Admin paneliga xush kelibsiz!", 
                             reply_markup=get_admin_menu(is_super))
    else:
        await bot.send_message(user_id, "👋 Salom! 'KITOBXON KIDS' botiga xush kelibsiz!", 
                             reply_markup=get_main_menu())
    
    await state.clear()

# 📋 Registration handlers
@dp.message(F.text == "📋 Ro'yxatdan o'tish")
async def register_start(message: types.Message, state: FSMContext):
    """Start registration process"""
    await message.answer("👶 Farzandingiz ism familiyasini kiriting:")
    await state.set_state(Registration.child_name)

@dp.message(Registration.child_name)
async def register_child_name(message: types.Message, state: FSMContext):
    """Register child name"""
    await state.update_data(child_name=message.text)
    await message.answer("👨‍👩‍👦 Ota-onaning ism familiyasini kiriting:")
    await state.set_state(Registration.parent_name)

@dp.message(Registration.parent_name)
async def register_parent_name(message: types.Message, state: FSMContext):
    """Register parent name"""
    await state.update_data(parent_name=message.text)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=region)] for region in REGIONS.keys()],
        resize_keyboard=True
    )
    await message.answer("🌍 Viloyatni tanlang:", reply_markup=keyboard)
    await state.set_state(Registration.region)

@dp.message(Registration.region)
async def register_region(message: types.Message, state: FSMContext):
    """Register region"""
    region = message.text
    if region not in REGIONS:
        await message.answer("❌ Iltimos, ro'yxatdan viloyat tanlang!")
        return
    
    await state.update_data(region=region)
    districts = REGIONS.get(region, [])
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=dist)] for dist in districts],
        resize_keyboard=True
    )
    await message.answer("🏙 Tumaningizni tanlang:", reply_markup=keyboard)
    await state.set_state(Registration.district)

@dp.message(Registration.district)
async def register_district(message: types.Message, state: FSMContext):
    """Register district"""
    await state.update_data(district=message.text)
    await message.answer("🏘 Mahallangiz nomini kiriting:")
    await state.set_state(Registration.mahalla)

@dp.message(Registration.mahalla)
async def register_mahalla(message: types.Message, state: FSMContext):
    """Register mahalla"""
    await state.update_data(mahalla=message.text)
    age_keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="7"), KeyboardButton(text="8"), KeyboardButton(text="9"), KeyboardButton(text="10")],
            [KeyboardButton(text="11"), KeyboardButton(text="12"), KeyboardButton(text="13"), KeyboardButton(text="14")]
        ],
        resize_keyboard=True
    )
    await message.answer("📅 Yoshni tanlang:", reply_markup=age_keyboard)
    await state.set_state(Registration.age)

@dp.message(Registration.age)
async def register_age(message: types.Message, state: FSMContext):
    """Register age"""
    if message.text not in ["7", "8", "9", "10", "11", "12", "13", "14"]:
        await message.answer("❌ Iltimos, 7-14 oralig'idagi yoshni tanlang!")
        return
    
    await state.update_data(age=message.text)
    phone_button = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📞 Telefon raqamni yuborish", request_contact=True)]],
        resize_keyboard=True
    )
    await message.answer("📞 Telefon raqamingizni yuboring:", reply_markup=phone_button)
    await state.set_state(Registration.phone)

@dp.message(Registration.phone)
async def register_phone(message: types.Message, state: FSMContext):
    """Register phone and complete registration"""
    if not message.contact:
        await message.answer("📞 Iltimos, tugma orqali telefon raqam yuboring.")
        return

    user_data = await state.get_data()
    phone_number = message.contact.phone_number
    user_data['phone'] = phone_number
    user_data['telegram_id'] = message.from_user.id
    user_data['username'] = message.from_user.username or "No username"
    user_data['registration_date'] = datetime.now().isoformat()

    # Save user data
    save_user(str(message.from_user.id), user_data)

    # Send registration info to all admins (both super admin and regular admins)
    reg_info = (
        f"📋 Yangi ro'yxatdan o'tish:\n"
        f"👶 Farzand: {user_data['child_name']}\n"
        f"👨‍👩‍👦 Ota-ona: {user_data['parent_name']}\n"
        f"🌍 Viloyat: {user_data['region']}\n"
        f"🏙 Tuman: {user_data['district']}\n"
        f"🏘 Mahalla: {user_data['mahalla']}\n"
        f"📅 Yosh: {user_data['age']}\n"
        f"📞 Telefon: {phone_number}\n"
        f"🆔 Telegram ID: {message.from_user.id}\n"
        f"👤 Username: @{user_data['username']}\n"
        f"📅 Ro'yxat sanasi: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )

    # Send to all admins including both super admin and regular admins
    admins = get_admins()
    for admin_id in admins.keys():
        try:
            await bot.send_message(int(admin_id), reg_info)
        except Exception as e:
            logging.error(f"Error sending message to admin {admin_id}: {e}")

    await message.answer("✅ Ro'yxatdan o'tish muvaffaqiyatli yakunlandi!", reply_markup=get_main_menu())
    await state.clear()

# 📝 Test handlers
@dp.message(F.text == "📝 Test topshirish")
async def start_test(message: types.Message, state: FSMContext):
    """Start test process"""
    user_id = str(message.from_user.id)
    users = get_users()
    
    if user_id not in users:
        await message.answer("❌ Test topshirishdan oldin ro'yxatdan o'ting!")
        return
    
    user_age = int(users[user_id]['age'])
    age_group = "7-10" if user_age <= 10 else "11-14"
    
    tests = get_tests()
    available_tests = tests.get(age_group, {})
    
    if not available_tests:
        await message.answer("❌ Sizning yosh guruhingiz uchun testlar mavjud emas!")
        return
    
    # Select 25 random questions from available tests
    all_questions = []
    for test_id, test_data in available_tests.items():
        questions = test_data.get('questions', [])
        all_questions.extend([(test_id, q) for q in questions])
    
    if len(all_questions) < 25:
        await message.answer("❌ Yetarli miqdorda test savollari mavjud emas!")
        return
    
    selected_questions = random.sample(all_questions, 25)
    
    # Initialize test session
    test_session = {
        'questions': selected_questions,
        'current_question': 0,
        'answers': [],
        'start_time': datetime.now().isoformat(),
        'age_group': age_group
    }
    
    await state.update_data(test_session=test_session)
    await send_next_question(message, state)

async def send_next_question(message: types.Message, state: FSMContext):
    """Send next test question"""
    data = await state.get_data()
    test_session = data['test_session']
    
    if test_session['current_question'] >= len(test_session['questions']):
        await complete_test(message, state)
        return
    
    question_data = test_session['questions'][test_session['current_question']][1]
    question_num = test_session['current_question'] + 1
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"A) {question_data['option_a']}", callback_data="answer_a")],
        [InlineKeyboardButton(text=f"B) {question_data['option_b']}", callback_data="answer_b")],
        [InlineKeyboardButton(text=f"C) {question_data['option_c']}", callback_data="answer_c")],
        [InlineKeyboardButton(text=f"D) {question_data['option_d']}", callback_data="answer_d")]
    ])
    
    question_text = f"📝 Savol {question_num}/25\n\n{question_data['question']}"
    
    await message.answer(question_text, reply_markup=keyboard)
    await state.set_state(TestStates.test_question)
    
    # Set 1-minute timer
    asyncio.create_task(question_timer(message.from_user.id, state, question_num))

async def question_timer(user_id: int, state: FSMContext, question_num: int):
    """1-minute timer for each question"""
    await asyncio.sleep(60)  # 1 minute
    
    try:
        current_state = await state.get_state()
        if current_state == TestStates.test_question:
            data = await state.get_data()
            test_session = data.get('test_session', {})
            
            if test_session.get('current_question', 0) + 1 == question_num:
                # Time's up for this question
                test_session['answers'].append({'answer': None, 'correct': False})
                test_session['current_question'] += 1
                await state.update_data(test_session=test_session)
                
                await bot.send_message(user_id, "⏰ Vaqt tugadi! Keyingi savolga o'tish...")
                
                # Send next question or complete test
                if test_session['current_question'] >= len(test_session['questions']):
                    await complete_test_by_id(user_id, state)
                else:
                    await send_next_question_by_id(user_id, state)
    except Exception as e:
        logging.error(f"Error in question timer: {e}")

async def send_next_question_by_id(user_id: int, state: FSMContext):
    """Send next question by user ID"""
    data = await state.get_data()
    test_session = data['test_session']
    
    question_data = test_session['questions'][test_session['current_question']][1]
    question_num = test_session['current_question'] + 1
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"A) {question_data['option_a']}", callback_data="answer_a")],
        [InlineKeyboardButton(text=f"B) {question_data['option_b']}", callback_data="answer_b")],
        [InlineKeyboardButton(text=f"C) {question_data['option_c']}", callback_data="answer_c")],
        [InlineKeyboardButton(text=f"D) {question_data['option_d']}", callback_data="answer_d")]
    ])
    
    question_text = f"📝 Savol {question_num}/25\n\n{question_data['question']}"
    
    await bot.send_message(user_id, question_text, reply_markup=keyboard)
    asyncio.create_task(question_timer(user_id, state, question_num))

async def complete_test_by_id(user_id: int, state: FSMContext):
    """Complete test by user ID"""
    data = await state.get_data()
    test_session = data['test_session']
    
    # Calculate results
    correct_answers = sum(1 for answer in test_session['answers'] if answer.get('correct', False))
    score = correct_answers * 4
    percentage = (correct_answers / 25) * 100
    
    start_time = datetime.fromisoformat(test_session['start_time'])
    end_time = datetime.now()
    time_taken = str(end_time - start_time).split('.')[0]
    
    # Get user data
    users = get_users()
    user_data = users.get(str(user_id), {})
    
    # Save result
    result_data = {
        'user_id': user_id,
        'user_name': user_data.get('child_name', 'N/A'),
        'telegram_id': user_id,
        'username': user_data.get('username', 'N/A'),
        'age': user_data.get('age', 'N/A'),
        'age_group': test_session['age_group'],
        'score': score,
        'correct_answers': correct_answers,
        'total_questions': 25,
        'percentage': round(percentage, 2),
        'time_taken': time_taken,
        'date': datetime.now().isoformat(),
        'answers': test_session['answers']
    }
    
    save_result(result_data)
    
    # Send result to user
    result_text = (
        f"✅ Test yakunlandi!\n\n"
        f"📊 Natijalar:\n"
        f"✅ To'g'ri javoblar: {correct_answers}/25\n"
        f"⭐ Ball: {score}/100\n"
        f"📈 Foiz: {percentage:.1f}%\n"
        f"⏱ Vaqt: {time_taken}"
    )
    
    await bot.send_message(user_id, result_text, reply_markup=get_main_menu())
    
    # Send result to admins
    admin_text = (
        f"📊 Yangi test natijasi:\n\n"
        f"👤 Foydalanuvchi: {user_data.get('child_name', 'N/A')}\n"
        f"🆔 Telegram ID: {user_id}\n"
        f"👤 Username: @{user_data.get('username', 'N/A')}\n"
        f"📅 Yosh: {user_data.get('age', 'N/A')}\n"
        f"✅ To'g'ri javoblar: {correct_answers}/25\n"
        f"⭐ Ball: {score}/100\n"
        f"📈 Foiz: {percentage:.1f}%\n"
        f"⏱ Vaqt: {time_taken}"
    )
    
    admins = get_admins()
    for admin_id in admins.keys():
        try:
            await bot.send_message(int(admin_id), admin_text)
        except Exception as e:
            logging.error(f"Error sending result to admin {admin_id}: {e}")
    
    await state.clear()

@dp.callback_query(F.data.in_(["answer_a", "answer_b", "answer_c", "answer_d"]))
async def handle_test_answer(callback_query: types.CallbackQuery, state: FSMContext):
    """Handle test answer"""
    data = await state.get_data()
    test_session = data.get('test_session', {})
    
    if not test_session:
        await callback_query.answer("❌ Test sessiyasi topilmadi!")
        return
    
    selected_answer = callback_query.data.split('_')[1]  # a, b, c, or d
    question_data = test_session['questions'][test_session['current_question']][1]
    correct_answer = question_data['correct_answer'].lower()
    
    is_correct = selected_answer == correct_answer
    test_session['answers'].append({'answer': selected_answer, 'correct': is_correct})
    test_session['current_question'] += 1
    
    await state.update_data(test_session=test_session)
    
    if is_correct:
        await callback_query.answer("✅ To'g'ri javob!")
    else:
        await callback_query.answer(f"❌ Noto'g'ri. To'g'ri javob: {correct_answer.upper()}")
    
    # Send next question or complete test
    if test_session['current_question'] >= len(test_session['questions']):
        await complete_test(callback_query.message, state)
    else:
        await send_next_question(callback_query.message, state)

async def complete_test(message: types.Message, state: FSMContext):
    """Complete test and show results"""
    data = await state.get_data()
    test_session = data['test_session']
    
    # Calculate results
    correct_answers = sum(1 for answer in test_session['answers'] if answer.get('correct', False))
    score = correct_answers * 4
    percentage = (correct_answers / 25) * 100
    
    start_time = datetime.fromisoformat(test_session['start_time'])
    end_time = datetime.now()
    time_taken = str(end_time - start_time).split('.')[0]
    
    # Get user data
    users = get_users()
    user_data = users.get(str(message.from_user.id), {})
    
    # Save result
    result_data = {
        'user_id': message.from_user.id,
        'user_name': user_data.get('child_name', 'N/A'),
        'telegram_id': message.from_user.id,
        'username': user_data.get('username', 'N/A'),
        'age': user_data.get('age', 'N/A'),
        'age_group': test_session['age_group'],
        'score': score,
        'correct_answers': correct_answers,
        'total_questions': 25,
        'percentage': round(percentage, 2),
        'time_taken': time_taken,
        'date': datetime.now().isoformat(),
        'answers': test_session['answers']
    }
    
    save_result(result_data)
    
    # Send result to user
    result_text = (
        f"✅ Test yakunlandi!\n\n"
        f"📊 Natijalar:\n"
        f"✅ To'g'ri javoblar: {correct_answers}/25\n"
        f"⭐ Ball: {score}/100\n"
        f"📈 Foiz: {percentage:.1f}%\n"
        f"⏱ Vaqt: {time_taken}"
    )
    
    await message.answer(result_text, reply_markup=get_main_menu())
    
    # Send result to admins
    admin_text = (
        f"📊 Yangi test natijasi:\n\n"
        f"👤 Foydalanuvchi: {user_data.get('child_name', 'N/A')}\n"
        f"🆔 Telegram ID: {message.from_user.id}\n"
        f"👤 Username: @{user_data.get('username', 'N/A')}\n"
        f"📅 Yosh: {user_data.get('age', 'N/A')}\n"
        f"✅ To'g'ri javoblar: {correct_answers}/25\n"
        f"⭐ Ball: {score}/100\n"
        f"📈 Foiz: {percentage:.1f}%\n"
        f"⏱ Vaqt: {time_taken}"
    )
    
    admins = get_admins()
    for admin_id in admins.keys():
        try:
            await bot.send_message(int(admin_id), admin_text)
        except Exception as e:
            logging.error(f"Error sending result to admin {admin_id}: {e}")
    
    await state.clear()

# 💬 Feedback handler
@dp.message(F.text == "💬 Fikr va maslahatlar")
async def feedback_prompt(message: types.Message, state: FSMContext):
    """Prompt for feedback"""
    await message.answer("✍️ Fikringizni yozing:")
    await state.set_state(Registration.feedback)

@dp.message(Registration.feedback)
async def save_feedback(message: types.Message, state: FSMContext):
    """Save feedback and send to all admins"""
    username = message.from_user.username or 'Username yoq'
    if username != 'Username yoq' and not username.startswith('@'):
        username = f"@{username}"
    
    full_name = message.from_user.full_name or 'Ism korsatilmagan'
    feedback_text = (
        f"💬 Yangi fikr va maslahat:\n\n"
        f"👤 Foydalanuvchi: {full_name}\n"
        f"🆔 Telegram ID: {message.from_user.id}\n"
        f"👤 Username: {username}\n"
        f"📅 Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"💭 Fikr: {message.text}"
    )
    
    # Send to all admins (both super admin and regular admins)
    admins = get_admins()
    for admin_id in admins.keys():
        try:
            await bot.send_message(int(admin_id), feedback_text)
        except Exception as e:
            logging.error(f"Error sending feedback to admin {admin_id}: {e}")
    
    await message.answer("✅ Fikringiz uchun rahmat!", reply_markup=get_main_menu())
    await state.clear()

# 📚 Project info handler
@dp.message(F.text == "📚 Loyiha haqida")
async def project_info(message: types.Message):
    """Show project information"""
    text = """<b>"Kitobxon kids" tanlovini tashkil etish va o'tkazish to'g'risidagi NIZOM</b>

🔹 <b>Umumiy qoidalar:</b>
• Mazkur Nizom yoshlar o'rtasida "Kitobxon Kids" tanlovini o'tkazish tartibini belgilaydi.
• Tanlov 7–10 va 11–14 yoshdagi bolalar uchun mo'ljallangan.
• Tanlov kitobxonlik madaniyatini oshirishga qaratilgan.

🔹 <b>Tashkilotchilar:</b>
• Yoshlar ishlari agentligi,
• Maktabgacha va maktab ta'limi vazirligi,
• O'zbekiston bolalar tashkiloti.

🔹 <b>Ishtirokchilar:</b>
• 7–14 yoshdagi barcha bolalar qatnasha oladi.
• Qoraqalpoq va rus tillarida ham qatnashish mumkin.

🔹 <b>Maqsad va vazifalar:</b>
• Kitob o'qishga qiziqish uyg'otish, mustaqil o'qish ko'nikmasini shakllantirish.
• Adiblar merosini o'rganish, o'zlikni anglashga chorlash.

🔹 <b>Tanlov bosqichlari:</b>
1. Saralash (oy boshida test, 25 ta savol, har biri 4 ball).
2. Hududiy (30 ta savol, har biri 30 soniya, top scorer keyingi bosqichga o'tadi).
3. Respublika (Fantaziya festivali, Taassurotlar, Savollar - 100 ballik tizim).

🔹 <b>G'oliblar:</b>
• 1-o'rin: Noutbuk
• 2-o'rin: Planshet
• 3-o'rin: Telefon
• Barcha qatnashchilarga velosiped

🔹 <b>Moliya manbalari:</b>
• Agentlik mablag'lari, homiylar, qonuniy xayriyalar.

Batafsil: @Kitobxon_Kids kanali orqali kuzatib boring.
"""
    await message.answer(text)

# 👨‍💼 Admin handlers
@dp.message(F.text == "🔙 Asosiy menyu")
async def back_to_main(message: types.Message, state: FSMContext):
    """Back to main menu"""
    await state.clear()
    if is_admin(message.from_user.id):
        is_super = is_super_admin(message.from_user.id)
        await message.answer("Admin panel", reply_markup=get_admin_menu(is_super))
    else:
        await message.answer("Asosiy menyu", reply_markup=get_main_menu())

@dp.message(F.text == "🔙 Orqaga")
async def back_button(message: types.Message, state: FSMContext):
    """Back button handler"""
    await state.clear()
    if is_admin(message.from_user.id):
        is_super = is_super_admin(message.from_user.id)
        await message.answer("Admin panel", reply_markup=get_admin_menu(is_super))
    else:
        await message.answer("Asosiy menyu", reply_markup=get_main_menu())

@dp.message(F.text == "👥 Foydalanuvchilar ro'yxati")
async def show_users(message: types.Message):
    """Show users list (admin only)"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    users = get_users()
    
    if not users:
        await message.answer("📝 Hozircha ro'yxatdan o'tgan foydalanuvchilar yo'q.")
        return
    
    users_text = "👥 Ro'yxatdan o'tgan foydalanuvchilar:\n\n"
    
    for i, (user_id, user_data) in enumerate(users.items(), 1):
        username = user_data.get('username', 'Username yoq')
        if username != 'Username yoq' and not username.startswith('@'):
            username = f"@{username}"
        
        users_text += (
            f"{i}. {user_data.get('child_name', 'N/A')}\n"
            f"   👤 Username: {username}\n"
            f"   🆔 Telegram ID: {user_id}\n"
            f"   📅 Yosh: {user_data.get('age', 'N/A')}\n"
            f"   🌍 Viloyat: {user_data.get('region', 'N/A')}\n"
            f"   🏙 Tuman: {user_data.get('district', 'N/A')}\n"
            f"   📞 Telefon: {user_data.get('phone', 'N/A')}\n"
            f"   📅 Ro'yxat sanasi: {user_data.get('registration_date', 'N/A')}\n\n"
        )
        
        # Split long messages
        if len(users_text) > 3500:
            await message.answer(users_text)
            users_text = ""
    
    if users_text:
        await message.answer(users_text)

@dp.message(F.text == "👨‍💼 Adminlar ro'yxati")
async def show_admins(message: types.Message):
    """Show admins list with full names and usernames (super admin only)"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    admins = get_admins()
    
    if not admins:
        await message.answer("📝 Adminlar ro'yxati bo'sh.")
        return
    
    admins_text = "👨‍💼 Adminlar ro'yxati:\n\n"
    
    for i, (admin_id, admin_data) in enumerate(admins.items(), 1):
        role = "🔴 Super Admin" if admin_data.get('role') == 'super_admin' else "🟡 Admin"
        added_date = admin_data.get('added_date', 'Noma\'lum')[:10] if admin_data.get('added_date') else 'Noma\'lum'
        
        # Try to get full name and username from Telegram
        try:
            admin_info = await bot.get_chat(int(admin_id))
            full_name = admin_info.full_name or 'Ism korsatilmagan'
            username = f"@{admin_info.username}" if admin_info.username else 'Username yoq'
        except Exception:
            full_name = 'Ism korsatilmagan'
            username = 'Username yoq'
        
        admins_text += f"{i}. {role}\n"
        admins_text += f"   👤 Ism: {full_name}\n"
        admins_text += f"   👤 Username: {username}\n"
        admins_text += f"   🆔 ID: {admin_id}\n"
        admins_text += f"   📅 Qo'shilgan: {added_date}\n\n"
        
        if len(admins_text) > 3000:  # Telegram message limit
            await message.answer(admins_text)
            admins_text = ""
    
    if admins_text.strip():
        await message.answer(admins_text)

@dp.message(F.text == "➕ Admin qo'shish")
async def add_admin_prompt(message: types.Message, state: FSMContext):
    """Prompt to add admin (super admin only)"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    await message.answer(
        "👨‍💼 Yangi admin qo'shish:\n\n"
        "Quyidagilardan birini kiriting:\n"
        "• Telegram ID (masalan: 123456789)\n"
        "• Username (masalan: @admin_user yoki admin_user)"
    )
    await state.set_state(AdminStates.add_admin)

@dp.message(AdminStates.add_admin)
async def add_admin(message: types.Message, state: FSMContext):
    """Add new admin by ID or username"""
    input_text = message.text.strip()
    admin_id = None
    admin_username = None
    
    # Check if input is a username
    if input_text.startswith('@'):
        admin_username = input_text[1:]  # Remove @
        # Try to get user info by username (this would require additional API calls)
        await message.answer("❌ Username orqali admin qo'shish hozircha ishlamaydi. Iltimos, Telegram ID kiriting.")
        return
    elif input_text.isalpha() or '_' in input_text:
        admin_username = input_text
        await message.answer("❌ Username orqali admin qo'shish hozircha ishlamaydi. Iltimos, Telegram ID kiriting.")
        return
    else:
        # Try to parse as ID
        try:
            admin_id = int(input_text)
        except ValueError:
            await message.answer("❌ Noto'g'ri format! Telegram ID (raqam) yoki username kiriting.")
            return
    
    if admin_id:
        admin_data = {
            'role': 'admin',
            'added_by': str(message.from_user.id),
            'added_date': datetime.now().isoformat(),
            'telegram_id': admin_id
        }
        
        save_admin(str(admin_id), admin_data)
        
        await message.answer(f"✅ Admin muvaffaqiyatli qo'shildi!\n🆔 Telegram ID: {admin_id}")
        
        # Notify new admin
        try:
            await bot.send_message(admin_id, 
                "🎉 Tabriklaymiz!\n\n"
                "Siz KITOBXON KIDS botida admin lavozimiga tayinlandingiz!\n"
                "Endi siz testlar qo'sha olasiz va foydalanuvchilar ro'yxatini ko'ra olasiz."
            )
        except Exception as e:
            logging.error(f"Error notifying new admin {admin_id}: {e}")
            await message.answer(f"⚠️ Admin qo'shildi, lekin xabar yuborishda xatolik: {e}")
    
    await state.clear()

@dp.message(F.text == "➖ Admin o'chirish")
async def remove_admin_prompt(message: types.Message, state: FSMContext):
    """Prompt to remove admin (super admin only)"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    admins = get_admins()
    regular_admins = {k: v for k, v in admins.items() if v.get('role') != 'super_admin'}
    
    if not regular_admins:
        await message.answer("📝 O'chiriladigan adminlar yo'q. Faqat oddiy adminlarni o'chirish mumkin.")
        return
    
    admin_list = "➖ Admin o'chirish:\n\n"
    admin_list += "O'chirish uchun admin ID kiriting:\n\n"
    
    for admin_id, admin_data in regular_admins.items():
        try:
            admin_info = await bot.get_chat(int(admin_id))
            full_name = admin_info.full_name or 'Ism korsatilmagan'
            username = f"@{admin_info.username}" if admin_info.username else 'Username yoq'
        except Exception:
            full_name = 'Ism korsatilmagan'
            username = 'Username yoq'
        
        admin_list += f"🟡 Admin\n"
        admin_list += f"   👤 Ism: {full_name}\n"
        admin_list += f"   👤 Username: {username}\n"
        admin_list += f"   🆔 ID: {admin_id}\n\n"
    
    await message.answer(admin_list)
    await message.answer("Admin ID kiriting:")
    await state.set_state(AdminStates.remove_admin)

@dp.message(AdminStates.remove_admin)
async def remove_admin_handler(message: types.Message, state: FSMContext):
    """Remove admin handler"""
    try:
        admin_id_to_remove = message.text.strip()
        
        # Check if the ID exists and is not a super admin
        admins = get_admins()
        if admin_id_to_remove not in admins:
            await message.answer("❌ Bunday ID li admin topilmadi!")
            return
        
        if admins[admin_id_to_remove].get('role') == 'super_admin':
            await message.answer("❌ Super adminni o'chirish mumkin emas!")
            return
        
        # Get admin info before removing
        try:
            admin_info = await bot.get_chat(int(admin_id_to_remove))
            full_name = admin_info.full_name or 'Ism korsatilmagan'
            username = f"@{admin_info.username}" if admin_info.username else 'Username yoq'
        except Exception:
            full_name = 'Ism korsatilmagan'
            username = 'Username yoq'
        
        # Remove admin
        if remove_admin(admin_id_to_remove):
            await message.answer(
                f"✅ Admin muvaffaqiyatli o'chirildi!\n\n"
                f"👤 Ism: {full_name}\n"
                f"👤 Username: {username}\n"
                f"🆔 ID: {admin_id_to_remove}"
            )
            
            # Notify the removed admin
            try:
                await bot.send_message(
                    int(admin_id_to_remove), 
                    "📢 Xabar!\n\n"
                    "Sizning KITOBXON KIDS botidagi admin huquqlaringiz bekor qilindi.\n"
                    "Endi siz oddiy foydalanuvchi sifatida botdan foydalanishingiz mumkin."
                )
            except Exception as e:
                logging.error(f"Error notifying removed admin: {e}")
        else:
            await message.answer("❌ Adminni o'chirishda xatolik yuz berdi!")
            
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
        logging.error(f"Error in remove_admin_handler: {e}")
    
    await state.clear()

@dp.message(F.text == "⬆️ Super Admin tayinlash")
async def promote_super_admin_prompt(message: types.Message, state: FSMContext):
    """Prompt to promote admin to super admin (super admin only)"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    admins = get_admins()
    regular_admins = {k: v for k, v in admins.items() if v.get('role') != 'super_admin'}
    
    if not regular_admins:
        await message.answer("📝 Super Admin tayinlanadigan oddiy adminlar yo'q.")
        return
    
    admin_list = "⬆️ Super Admin tayinlash:\n\n"
    admin_list += "Super Admin qilish uchun admin ID kiriting:\n\n"
    
    for admin_id, admin_data in regular_admins.items():
        try:
            admin_info = await bot.get_chat(int(admin_id))
            full_name = admin_info.full_name or 'Ism korsatilmagan'
            username = f"@{admin_info.username}" if admin_info.username else 'Username yoq'
        except Exception:
            full_name = 'Ism korsatilmagan'
            username = 'Username yoq'
        
        admin_list += f"🟡 Admin\n"
        admin_list += f"   👤 Ism: {full_name}\n"
        admin_list += f"   👤 Username: {username}\n"
        admin_list += f"   🆔 ID: {admin_id}\n\n"
    
    await message.answer(admin_list)
    await message.answer("Super Admin qilmoqchi bo'lgan admin ID sini kiriting:")
    await state.set_state(AdminStates.promote_super_admin)

@dp.message(AdminStates.promote_super_admin)
async def promote_super_admin_handler(message: types.Message, state: FSMContext):
    """Promote admin to super admin handler"""
    try:
        admin_id_to_promote = message.text.strip()
        
        # Check if the ID exists and is a regular admin
        admins = get_admins()
        if admin_id_to_promote not in admins:
            await message.answer("❌ Bunday ID li admin topilmadi!")
            return
        
        if admins[admin_id_to_promote].get('role') == 'super_admin':
            await message.answer("❌ Bu admin allaqachon Super Admin!")
            return
        
        # Get admin info before promoting
        try:
            admin_info = await bot.get_chat(int(admin_id_to_promote))
            full_name = admin_info.full_name or 'Ism korsatilmagan'
            username = f"@{admin_info.username}" if admin_info.username else 'Username yoq'
        except Exception:
            full_name = 'Ism korsatilmagan'
            username = 'Username yoq'
        
        # Update admin role to super_admin
        admin_data = admins[admin_id_to_promote].copy()
        admin_data['role'] = 'super_admin'
        admin_data['promoted_by'] = str(message.from_user.id)
        admin_data['promoted_date'] = datetime.now().isoformat()
        
        save_admin(admin_id_to_promote, admin_data)
        
        await message.answer(
            f"✅ Super Admin muvaffaqiyatli tayinlandi!\n\n"
            f"👤 Ism: {full_name}\n"
            f"👤 Username: {username}\n"
            f"🆔 ID: {admin_id_to_promote}\n\n"
            f"🔴 Endi bu foydalanuvchi Super Admin huquqlariga ega!"
        )
        
        # Notify the promoted admin
        try:
            await bot.send_message(
                int(admin_id_to_promote), 
                "🎉 Tabriklaymiz!\n\n"
                "Siz KITOBXON KIDS botida Super Admin lavozimiga tayinlandingiz!\n\n"
                "🔴 Endi sizda quyidagi huquqlar bor:\n"
                "• Adminlarni qo'shish va o'chirish\n"
                "• Super Admin tayinlash\n"
                "• Testlarni qo'shish va o'chirish\n"
                "• Barcha hisobotlarni yuklab olish\n"
                "• Barcha ma'lumotlarni ko'rish\n\n"
                "Mas'uliyat bilan foydalaning!"
            )
        except Exception as e:
            logging.error(f"Error notifying promoted super admin: {e}")
            
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
        logging.error(f"Error in promote_super_admin_handler: {e}")
    
    await state.clear()

@dp.message(F.text == "➕ Test qo'shish")
async def add_test_age_prompt(message: types.Message, state: FSMContext):
    """Prompt to select age group for new test"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    await message.answer("📅 Yosh guruhini tanlang:", reply_markup=get_age_group_keyboard())
    await state.set_state(AdminStates.add_test_age)

@dp.message(AdminStates.add_test_age)
async def add_test_age(message: types.Message, state: FSMContext):
    """Set age group for test"""
    if message.text == "🔙 Orqaga":
        await back_button(message, state)
        return
    
    if message.text not in ["7-10 yosh", "11-14 yosh"]:
        await message.answer("❌ Iltimos, yosh guruhini tanlang!")
        return
    
    age_group = message.text.split()[0]  # "7-10" or "11-14"
    await state.update_data(age_group=age_group)
    await message.answer("📚 Kitob nomini kiriting:")
    await state.set_state(AdminStates.add_test_book)

@dp.message(AdminStates.add_test_book)
async def add_test_book(message: types.Message, state: FSMContext):
    """Set book name for test"""
    await state.update_data(book_name=message.text)
    
    content_keyboard = ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📝 Matn ko'rinishida")],
        [KeyboardButton(text="📄 PDF fayl ko'rinishida")],
        [KeyboardButton(text="🔙 Orqaga")]
    ], resize_keyboard=True)
    
    await message.answer("📋 Test qanday formatda qo'shmoqchisiz?", reply_markup=content_keyboard)
    await state.set_state(AdminStates.add_test_content)

@dp.message(AdminStates.add_test_content)
async def add_test_content_type(message: types.Message, state: FSMContext):
    """Handle test content type selection"""
    if message.text == "🔙 Orqaga":
        await back_button(message, state)
        return
    
    if message.text == "📝 Matn ko'rinishida":
        await state.update_data(content_type="text")
        await message.answer(
            "📝 Testlarni quyidagi formatda kiriting:\n\n"
            "Savol matni?\n"
            "A) Variant 1\n"
            "B) Variant 2\n"
            "C) Variant 3\n"
            "D) Variant 4\n"
            "Javob: A\n\n"
            "Har bir test yangi qatordan boshlansin. Barcha testlarni bir xabarada yuboring."
        )
        await state.set_state(AdminStates.add_test_questions)
    
    elif message.text == "📄 PDF fayl ko'rinishida":
        await state.update_data(content_type="pdf")
        await message.answer("📄 PDF faylni yuboring:")
        await state.set_state(AdminStates.add_test_questions)
    else:
        await message.answer("❌ Iltimos, formatni tanlang!")

@dp.message(AdminStates.add_test_questions)
async def add_test_questions(message: types.Message, state: FSMContext):
    """Process test questions"""
    data = await state.get_data()
    content_type = data.get('content_type')
    
    questions = []
    
    if content_type == "text" and message.text:
        # Parse text format questions
        questions = parse_text_questions(message.text)
    
    elif content_type == "pdf" and message.document:
        # Handle PDF file
        if message.document.mime_type == "application/pdf":
            try:
                file = await bot.download(message.document)
                text_content = extract_text_from_pdf(file.read())
                questions = parse_text_questions(text_content)
            except Exception as e:
                await message.answer(f"❌ PDF ni o'qishda xatolik: {e}")
                return
        else:
            await message.answer("❌ Iltimos, PDF fayl yuboring!")
            return
    else:
        await message.answer("❌ Noto'g'ri format!")
        return
    
    if not questions:
        await message.answer("❌ Testlar topilmadi yoki noto'g'ri format!")
        return
    
    # Save test
    test_data = {
        'book_name': data['book_name'],
        'age_group': data['age_group'],
        'questions': questions,
        'added_by': str(message.from_user.id),
        'added_date': datetime.now().isoformat(),
        'content_type': content_type
    }
    
    save_test(test_data)
    
    is_super = is_super_admin(message.from_user.id)
    await message.answer(
        f"✅ Test muvaffaqiyatli qo'shildi!\n"
        f"📚 Kitob: {data['book_name']}\n"
        f"📅 Yosh guruhi: {data['age_group']}\n"
        f"📝 Savollar soni: {len(questions)}",
        reply_markup=get_admin_menu(is_super)
    )
    
    await state.clear()

def parse_text_questions(text: str) -> List[Dict]:
    """Parse questions from text format"""
    questions = []
    
    # Split by double newlines to separate questions
    question_blocks = text.strip().split('\n\n')
    
    for block in question_blocks:
        lines = [line.strip() for line in block.split('\n') if line.strip()]
        
        if len(lines) < 6:  # Need at least question + 4 options + answer
            continue
        
        question_text = lines[0].rstrip('?') + '?'
        
        options = {}
        correct_answer = None
        
        for line in lines[1:]:
            if line.startswith('A)') or line.startswith('А)'):
                options['option_a'] = line[2:].strip()
            elif line.startswith('B)') or line.startswith('Б)'):
                options['option_b'] = line[2:].strip()
            elif line.startswith('C)') or line.startswith('В)'):
                options['option_c'] = line[2:].strip()
            elif line.startswith('D)') or line.startswith('Г)'):
                options['option_d'] = line[2:].strip()
            elif line.lower().startswith('javob:') or line.lower().startswith('ответ:'):
                answer_part = line.split(':', 1)[1].strip().upper()
                if answer_part in ['A', 'B', 'C', 'D', 'А', 'Б', 'В', 'Г']:
                    # Convert Cyrillic to Latin if needed
                    if answer_part == 'А':
                        correct_answer = 'A'
                    elif answer_part == 'Б':
                        correct_answer = 'B'
                    elif answer_part == 'В':
                        correct_answer = 'C'
                    elif answer_part == 'Г':
                        correct_answer = 'D'
                    else:
                        correct_answer = answer_part
        
        if len(options) == 4 and correct_answer:
            questions.append({
                'question': question_text,
                'option_a': options['option_a'],
                'option_b': options['option_b'],
                'option_c': options['option_c'],
                'option_d': options['option_d'],
                'correct_answer': correct_answer
            })
    
    return questions

@dp.message(F.text == "🗑 Test o'chirish")
async def delete_test_age_prompt(message: types.Message, state: FSMContext):
    """Prompt to select age group for deleting test (super admin only)"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    await message.answer("📅 Qaysi yosh guruhidan test o'chirmoqchisiz?", reply_markup=get_age_group_keyboard())
    await state.set_state(AdminStates.delete_test_age)

@dp.message(AdminStates.delete_test_age)
async def delete_test_age(message: types.Message, state: FSMContext):
    """Select age group for test deletion"""
    if message.text == "🔙 Orqaga":
        await back_button(message, state)
        return
    
    if message.text not in ["7-10 yosh", "11-14 yosh"]:
        await message.answer("❌ Iltimos, yosh guruhini tanlang!")
        return
    
    age_group = message.text.split()[0]  # "7-10" or "11-14"
    tests = get_tests()
    available_tests = tests.get(age_group, {})
    
    if not available_tests:
        await message.answer(f"❌ {age_group} yosh guruhi uchun testlar mavjud emas!")
        await state.clear()
        return
    
    # Create keyboard with available tests
    keyboard = []
    for test_id, test_data in available_tests.items():
        book_name = test_data.get('book_name', 'Noma\'lum kitob')
        questions_count = len(test_data.get('questions', []))
        keyboard.append([KeyboardButton(text=f"{book_name} ({questions_count} savol) - {test_id[:8]}...")])
    
    keyboard.append([KeyboardButton(text="🔙 Orqaga")])
    
    test_keyboard = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    await state.update_data(age_group=age_group, available_tests=available_tests)
    await message.answer("🗑 O'chirish uchun testni tanlang:", reply_markup=test_keyboard)
    await state.set_state(AdminStates.delete_test_select)

@dp.message(AdminStates.delete_test_select)
async def delete_test_confirm(message: types.Message, state: FSMContext):
    """Confirm test deletion"""
    if message.text == "🔙 Orqaga":
        await back_button(message, state)
        return
    
    data = await state.get_data()
    age_group = data['age_group']
    available_tests = data['available_tests']
    
    # Find selected test by partial ID
    selected_test_id = None
    for test_id in available_tests.keys():
        if test_id[:8] in message.text:
            selected_test_id = test_id
            break
    
    if not selected_test_id:
        await message.answer("❌ Test topilmadi!")
        return
    
    # Delete test
    tests = get_tests()
    del tests[age_group][selected_test_id]
    save_json_data(TESTS_FILE, tests)
    
    test_name = available_tests[selected_test_id].get('book_name', 'Noma\'lum')
    
    await message.answer(
        f"✅ Test muvaffaqiyatli o'chirildi!\n📚 Kitob: {test_name}",
        reply_markup=get_admin_menu(True)
    )
    
    await state.clear()

@dp.message(F.text == "📊 Test natijalarini yuklab olish")
async def download_test_results(message: types.Message):
    """Download test results in PDF and Excel format (super admin only)"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    results = get_results()
    
    if not results:
        await message.answer("📊 Hozircha test natijalari mavjud emas.")
        return
    
    try:
        await message.answer("📊 Test natijalari tayyorlanmoqda...")
        
        # Generate PDF report
        pdf_data = generate_pdf_report(results)
        pdf_file = BufferedInputFile(pdf_data, filename="test_results.pdf")
        
        # Generate Excel report
        excel_data = generate_excel_report(results)
        excel_file = BufferedInputFile(excel_data, filename="test_results.xlsx")
        
        # Send PDF
        await bot.send_document(
            message.from_user.id,
            pdf_file,
            caption="📄 Test natijalari (PDF format)"
        )
        
        # Send Excel
        await bot.send_document(
            message.from_user.id,
            excel_file,
            caption="📊 Test natijalari (Excel format)"
        )
        
        await message.answer("✅ Test natijalari muvaffaqiyatli yuklandi!")
        
    except Exception as e:
        logging.error(f"Error generating test results reports: {e}")
        await message.answer(f"❌ Test natijalari hisobotini yaratishda xatolik: {e}")

@dp.message(F.text == "📋 Foydalanuvchi ma'lumotlarini yuklab olish")
async def download_user_data(message: types.Message):
    """Download user data in PDF and Excel format (super admin only)"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Sizda bu buyruqni bajarish huquqi yo'q!")
        return
    
    users = get_users()
    
    if not users:
        await message.answer("📋 Hozircha ro'yxatdan o'tgan foydalanuvchilar yo'q.")
        return
    
    try:
        await message.answer("📋 Foydalanuvchi ma'lumotlari tayyorlanmoqda...")
        
        # Generate PDF report
        pdf_data = generate_users_pdf_report(users)
        pdf_file = BufferedInputFile(pdf_data, filename="users_data.pdf")
        
        # Generate Excel report
        excel_data = generate_users_excel_report(users)
        excel_file = BufferedInputFile(excel_data, filename="users_data.xlsx")
        
        # Send PDF
        await bot.send_document(
            message.from_user.id,
            pdf_file,
            caption="📄 Foydalanuvchi ma'lumotlari (PDF format)"
        )
        
        # Send Excel
        await bot.send_document(
            message.from_user.id,
            excel_file,
            caption="📊 Foydalanuvchi ma'lumotlari (Excel format)"
        )
        
        await message.answer("✅ Foydalanuvchi ma'lumotlari muvaffaqiyatli yuklandi!")
        
    except Exception as e:
        logging.error(f"Error generating user data reports: {e}")
        await message.answer(f"❌ Foydalanuvchi ma'lumotlari hisobotini yaratishda xatolik: {e}")

# 🛡 Security handlers
@dp.message(F.text.contains("t.me") | F.text.contains("http") | F.text.contains("@"))
async def block_ads(message: types.Message):
    """Block advertisements and spam"""
    if message.chat.type == "private":
        try:
            await message.delete()
        except Exception:
            pass

# 📣 Main function
async def main():
    """Main function to start the bot"""
    await bot.delete_webhook(drop_pending_updates=True)
    
    # Initialize data files
    if not os.path.exists(ADMINS_FILE):
        default_admins = {str(SUPER_ADMIN_ID): {"role": "super_admin", "added_by": "system", "added_date": datetime.now().isoformat()}}
        save_json_data(ADMINS_FILE, default_admins)
    
    if not os.path.exists(TESTS_FILE):
        save_json_data(TESTS_FILE, {"7-10": {}, "11-14": {}})
    
    if not os.path.exists(USERS_FILE):
        save_json_data(USERS_FILE, {})
    
    if not os.path.exists(RESULTS_FILE):
        save_json_data(RESULTS_FILE, [])
    
    logging.info("Bot started successfully!")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
